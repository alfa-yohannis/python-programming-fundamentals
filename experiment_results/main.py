import os
import re
from openpyxl import load_workbook

# List of Excel files
files = [
    "courseid_4038_participants.xlsx",
    # "courseid_4042_participants.xlsx",
]

OUTPUT_DIR = "results"
os.makedirs(OUTPUT_DIR, exist_ok=True)

def extract_student_id(raw_first_name: str) -> str:
    """Extract numeric ID from 'First name' (e.g., '2510101054 -')."""
    if not isinstance(raw_first_name, str):
        return ""
    match = re.search(r"(\d+)", raw_first_name)
    return match.group(1) if match else ""

def clean_name(raw_last_name: str) -> str:
    """Normalize name for a filesystem-safe folder."""
    if not isinstance(raw_last_name, str):
        return ""
    name = re.sub(r"\s+", "_", raw_last_name.strip())
    name = re.sub(r"[^A-Za-z0-9_\-]", "", name)
    return name

for file in files:
    if not os.path.exists(file):
        print(f"⚠️ File not found: {file}")
        continue

    wb = load_workbook(filename=file, data_only=True)
    sheet = wb.active

    # Find header columns
    headers = [cell.value for cell in sheet[1]]
    try:
        first_col = headers.index("First name") + 1
        last_col = headers.index("Last name") + 1
    except ValueError:
        print(f"⚠️ Columns not found in {file}, skipping.")
        continue

    for row in sheet.iter_rows(min_row=2, values_only=True):
        first_name = row[first_col - 1]
        last_name = row[last_col - 1]

        sid = extract_student_id(str(first_name))
        name = clean_name(str(last_name))

        # Skip instructors or malformed rows
        if not sid or not name:
            continue

        folder_name = f"{sid}_{name}"
        student_path = os.path.join(OUTPUT_DIR, folder_name)

        if os.path.exists(student_path):
            print(f"⏭️  Skipping existing folder: {folder_name}")
        else:
            os.makedirs(student_path)
            print(f"✅ Created: {folder_name}")

        # Create soal_1 ... soal_7 under each student folder
        for i in range(1, 8):
            soal_path = os.path.join(student_path, f"soal_{i}")
            if not os.path.exists(soal_path):
                os.makedirs(soal_path)
                print(f"   └── 📁 Created {os.path.basename(soal_path)}")

print("✅ All folders created successfully.")
